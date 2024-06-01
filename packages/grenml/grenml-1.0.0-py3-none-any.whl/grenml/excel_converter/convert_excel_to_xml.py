"""
Copyright 2020 GRENMap Authors

SPDX-License-Identifier: Apache License 2.0

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.

--------------------------------------------------------------------

Synopsis: Take in a provided XLS workbook and converts it into GRENML.
    This script has been writen for GRENML 0.1.5.
    Later versions might not be compatible.
    Example use: python3 ./excel_conversion.py
        -o MyISP\ Inc.
        -t MyISP\ GRENML\ Topology
        -r ./MyISP\ GREN\ Map\ Data\ Form.xlsx
        -w ./output/MyISP_GRENML.xml

"""  # noqa: W605

from functools import reduce

from openpyxl import load_workbook
from openpyxl.cell import ReadOnlyCell
from grenml import GRENMLManager
from grenml.exceptions import (
    InstitutionNotFoundError, MultipleReturnedError, NodeNotFoundError,
    AttributeIdError, AttributeNameError
)

DEFAULT_TEMPLATE_VERSION = 'v0.0'

WORKSHEET_LOCATIONS = {

    # Sheets without a version code
    DEFAULT_TEMPLATE_VERSION: {

        'sheet_names': {
            'institutions': 'Institutions',
            'nodes': 'Nodes',
            'links': 'Links',
        },

        'first_row': {
            'institutions': 6,
            'nodes': 6,
            'links': 6,
        },

        # The list of fields to correlate the XLSX fields to the
        # Institution fields
        # This list doubles for iteration and for positional reference
        'institution_fields': {
            0: 'name',
            1: 'latitude',
            2: 'longitude',
            3: 'altitude',
            4: 'address',
            5: 'url',
            6: 'tag',
        },

        # The list of fields to correlate the XLSX fields to the Node
        # fields
        'node_fields': {
            0: 'name',
            1: 'short_name',
            2: 'id',
            3: 'latitude',
            4: 'longitude',
            5: 'altitude',
            6: 'address',
            7: 'tag',
            8: 'owners',  # There are multiple owners listed because in the model
            9: 'owners',  # owners is a list of IDs and 2 columns correlate to owners in the
                          # spreadsheet
        },

        # The list of fields to correlate the XLSX fields to the Link
        # fields
        'link_fields': {
            0: 'name',
            1: 'short_name',
            2: 'id',
            3: 'nodes',  # There are multiple nodes listed because the link model
            4: 'nodes',  # does not maintain direction between node.
            5: 'tag',
            6: 'throughput',
            7: 'owners',
            8: 'owners',
        },

    },

    'v1.0': {

        'sheet_names': {
            'institutions': 'Institutions',
            'nodes': 'Nodes',
            'links': 'Links',
        },

        'first_row': {
            'institutions': 7,
            'nodes': 6,
            'links': 6,
        },

        # The list of fields to correlate the XLSX fields to the
        # Institution fields
        # This list doubles for iteration and for positional reference
        'institution_fields': {
            0: 'name',
            1: 'latitude',
            2: 'longitude',
            3: 'altitude',
            4: 'address',
            5: 'url',
            6: 'description',
            7: 'tag',
        },

        # The list of fields to correlate the XLSX fields to the Node
        # fields
        'node_fields': {
            0: 'name',
            1: 'short_name',
            2: 'id',
            3: 'latitude',
            4: 'longitude',
            5: 'altitude',
            6: 'address',
            7: 'description',
            8: 'tag',
            9: 'owners',  # There are multiple owners listed because in the model
            10: 'owners',  # owners is a list of IDs and 2 columns correlate to owners in the
                           # spreadsheet
        },

        # The list of fields to correlate the XLSX fields to the Link
        # fields
        'link_fields': {
            0: 'name',
            1: 'short_name',
            2: 'id',
            3: 'nodes',  # There are multiple nodes listed because the link model
            4: 'nodes',  # does not maintain direction between node.
            5: 'description',
            6: 'tag',
            7: 'throughput',
            8: 'owners',
            9: 'owners',
        },

    },

}


def concat_owners(owner, owners):
    return [item.strip() for item in [str(owner)] + str(owners).split(';')]


def empty_row(row):
    row = tuple(item if not isinstance(item, ReadOnlyCell) else None for item in row)
    for item in row:
        if item:
            return row
    return None


class ExcelParseError(Exception):
    """
    XLSParser.parse_file attempts to continue reading a spreadsheet
    when it encounters problems like references to missing institutions,
    nodes with IDs that have already been registered and invalid
    geographical coordinates.

    The parse_file method uses a list to keep track of these problems.
    If any are found, it raises this exception to deliver the list
    to its caller.
    """
    def __init__(self, error_list):
        super().__init__()
        self.error_list = error_list


# Dictionary of templates for excel conversion error messages.
EXCEPTION_MESSAGES = {
    ValueError: (
        '{type} "{name}" has invalid latitude or longitude.'
    ),
    InstitutionNotFoundError: (
        'Institution not found: "{name}". '
        'Referenced by {reference_owner_type}: "{reference_owner_name}"'
    ),
    MultipleReturnedError: (
        '{type} "{name}" has multiple instances. '
        'Referenced by {reference_owner_type}: "{reference_owner_name}"'
    ),
    AttributeNameError: '{type} doesn\'t have Name defined.',
    AttributeIdError: '{type} "{name}" has duplicate values.',
    NodeNotFoundError: (
        'Node not found: "{name}". '
        'Referenced by link: "{reference_owner_name}"'
    )
}


def append_error_message(error_message_list, error):
    """
    Takes a list of messages (error_message_list)
    and an exception instance (error).

    Creates an error message from the exception and appends it
    to the list.

    Returns the list with the new error message.
    """
    error_type = type(error)
    if error_type in EXCEPTION_MESSAGES:
        error_data = error.args[0]
        message = EXCEPTION_MESSAGES[error_type].format(**error_data)
    else:
        message = repr(error)
    error_message_list.append(message)
    return error_message_list


class XLSParser:
    def __init__(self, topology_name, owner):
        self.manager = GRENMLManager(name=topology_name)
        self.owner = owner

    def add_institution(self, **institution_fields):
        """
        Asks the GRENMLManager to create an institution
        for its current topology.
        Intercepts ValueErrors due to invalid latitude or longitude.
        """
        result = None
        try:
            self.manager.add_institution(**institution_fields)
            result = {'success': True}
        except MultipleReturnedError:
            institutions = self.manager.get_institutions(name=institution_fields['name'])
            inst = institutions[0]
            result = {
                'success': False,
                'error': MultipleReturnedError({
                    'type': 'Institution',
                    'name': inst.name,
                    'reference_owner_type': inst.reference_owner_type,
                    'reference_owner_name': inst.reference_owner_name
                }),
                'value': list(institutions)[0] if institutions is not None else []
            }
        except AttributeNameError:
            result = {
                'success': False,
                'error': AttributeNameError({
                    'type': 'Institution'
                })
            }
        except ValueError:
            result = {
                'success': False,
                'error': ValueError({'type': 'Institution', 'name': institution_fields['name']})
            }
        except AttributeIdError:
            institutions = self.manager.get_institutions(name=institution_fields['name'])
            result = {
                'success': False,
                'error': AttributeIdError({
                    'type': 'Institution',
                    'name': institution_fields['name']
                }),
                'value': list(institutions)[0] if institutions is not None else []
            }
        return result

    def get_institution_by_name(self, name, reference_owner_type, reference_owner_name):
        """
        Requests an institution by name from GRENMLManager.
        Intercepts the "not found" and "multiple returned" errors.
        """
        name = str(name).strip()
        result = None
        try:
            value = self.manager.get_institution(name=name)
            result = {
                'success': True,
                'value': value
            }
        except InstitutionNotFoundError:
            value = self.manager.add_institution(name=name)
            result = {
                'success': False,
                'error': InstitutionNotFoundError({
                    'name': name,
                    'reference_owner_type': reference_owner_type,
                    'reference_owner_name': reference_owner_name
                }),
                'value': value
            }
        except MultipleReturnedError:
            institutions = self.manager.get_institutions(name=name)
            result = {
                'success': False,
                'error': MultipleReturnedError({
                    'type': 'institution',
                    'name': name,
                    'reference_owner_type': reference_owner_type,
                    'reference_owner_name': reference_owner_name
                }),
                'value': list(institutions)[0] if institutions is not None else []
            }
        return result

    def add_node(self, **node_fields):
        """
        Asks GRENMLManager to create a node.
        Intercepts ValueErrors for invalid latitude or longitude and
        AttributeIdErrors, which will occur if a node already exists,
        whose id is equal to the one provided for the new node.
        """
        result = None
        try:
            if 'name' in node_fields:
                value = self.manager.add_node(**node_fields)
                result = {
                    'success': True,
                    'value': value
                }
            else:
                result = {
                    'success': False,
                    'error': AttributeNameError({'type': 'Node'})
                }
        except ValueError:
            result = {
                'success': False,
                'error': ValueError({'type': 'Node', 'name': node_fields['name']})
            }
        except AttributeNameError:
            result = {
                'success': False,
                'error': AttributeNameError({
                    'type': 'Node'
                })
            }
        except AttributeIdError:
            nodes = self.manager.get_nodes(name=node_fields['name'])
            result = {
                'success': False,
                'error': AttributeIdError({'type': 'Node', 'name': node_fields['name']}),
                'value': list(nodes)[0] if nodes is not None else []
            }
        return result

    def get_node_by_name(self, name, link_name):
        """
        Requests a node by name from GRENMLManager.
        Handles the "not found" error by creating a placeholder node.
        Handles also the "multiple returned" error by selecting
        the first instance in the list of nodes with the given name.
        """
        result = None
        try:
            value = self.manager.get_node(name=name)
            result = {
                'success': True,
                'value': value
            }
        except NodeNotFoundError:
            self.manager.add_node(name=name)
            value = self.manager.get_node(name=name)
            result = {
                'success': False,
                'error': NodeNotFoundError({
                    'name': name,
                    'reference_owner_name': link_name
                }),
                'value': value
            }
        except MultipleReturnedError:
            nodes = self.manager.get_nodes(name=name)
            result = {
                'success': False,
                'error': MultipleReturnedError({
                    'type': 'node',
                    'name': name,
                    'reference_owner_type': 'link',
                    'reference_owner_name': link_name
                }),
                'value': list(nodes)[0] if nodes is not None else []
            }
        return result

    def add_link(self, **link_fields):
        """
        Asks GRENMLManager to create a link.
        Intercepts ValueErrors for invalid latitude or longitude and
        AttributeIdErrors, which will occur if a link already exists,
        whose id is equal to the one provided for the new link.
        """
        result = None
        try:
            if 'name' in link_fields:
                value = self.manager.add_link(**link_fields)
                result = {
                    'success': True,
                    'value': value
                }
            else:
                result = {
                    'success': False,
                    'error': AttributeNameError({'type': 'Link'})
                }
        except ValueError:
            result = {
                'success': False,
                'error': ValueError({'type': 'Link', 'name': link_fields['name']})
            }
        except AttributeNameError:
            result = {
                'success': False,
                'error': AttributeNameError({
                    'type': 'Link'
                })
            }
        except MultipleReturnedError:
            links = self.manager.get_links(name=link_fields['name'])
            link = links[0]
            result = {
                'success': False,
                'error': MultipleReturnedError({
                    'type': 'Link',
                    'name': link_fields['name'],
                    'reference_owner_type': 'link',
                    'reference_owner_name': link.link_name
                }),
                'value': list(links)[0] if links is not None else []
            }
        except AttributeIdError:
            # Modify the id received so the add_link call succeeds.
            links = self.manager.get_links(name=link_fields['name'])
            result = {
                'success': False,
                'error': AttributeIdError({'type': 'Link', 'name': link_fields['name']}),
                'value': list(links)[0] if links is not None else []
            }
        return result

    def parse_file(self, filename):
        """
        Reads an XSLX file and imports its data into a
        GRENMLManager object.

        Raises an ExcelParseError in case there are
        missing or duplicate institutions or nodes,
        multiple nodes with the same id,
        nodes or institutions with incorrectly formatted
        geographical coordinates.
        """
        manager, errors = self.parse_workbook(
            load_workbook(filename, read_only=True),
        )
        if errors:
            errors = reduce(append_error_message, errors, [])
        return (manager, errors)

    def parse_workbook(self, workbook):
        """
        Imports the data from an XLSX worbook into a
        GRENMLManager object.

        As the data is imported into the GRENMLManager,
        it is done in an order of Institution, Nodes, then Links.
        This is being performed in this way to allow for the
        GRENML objects to be created and the IDs for linking
        the objects together have been made.

        Every object has mandatory fields, like names,
        that must be filled for the conversion to GRENML to work and
        optional fields to allow for providing more information.

        :param filename: The file location of the XLSX file
        :return: A tuple in which the first element is
        a GRENMLManager object with all the data in the XLSX workbook
        and the second is a list of dictionaries representing errors
        found in the workbook.
        """
        errors = []

        # Get the version code for this spreadsheet template, always
        # assumed to be located
        # on the Directions sheet at B2, if that sheet exists
        template_version = DEFAULT_TEMPLATE_VERSION

        if 'Directions' in workbook:
            directions_sheet = workbook['Directions']
            if directions_sheet['B2'].value:
                template_version = directions_sheet['B2'].value

        print('Detected spreadsheet template version {}'.format(template_version))

        # Collect the institutions in the Workbook
        institution_sheet = workbook[
            WORKSHEET_LOCATIONS[template_version]['sheet_names']['institutions']]
        primary_owner_imported = False
        for row in institution_sheet.iter_rows(
            min_row=WORKSHEET_LOCATIONS[template_version]['first_row']['institutions'],
            max_col=(max(WORKSHEET_LOCATIONS[template_version]['institution_fields'].keys()) + 1),
            values_only=True,
        ):
            # Skipping over empty rows
            row = empty_row(row)
            if not row:
                continue

            print('INSTITUTION: {}'.format(row[0]))

            institution_fields = {
                'name': None
            }

            # Collect all fields
            for column, key in WORKSHEET_LOCATIONS[template_version]['institution_fields'].items():
                # RYAN: turn this into a function, and skip empty
                # entries in semicolon-delimited lists (from e.g.
                # trailing semicolons)
                if row[column]:
                    if ';' in str(row[column]):
                        institution_fields[key] = [
                            item.strip() for item in str(row[column]).split(';')]
                    else:
                        try:
                            if key == 'name' and row[column].strip() == '':
                                raise AttributeNameError
                            institution_fields[key] = row[column].strip()
                        except AttributeError:
                            # Likely not a string; just capture the
                            # value as-is
                            # errors.append(
                            #   AttributeNameError(
                            #       {'type': 'Institution'}))
                            institution_fields[key] = row[column]
                        except AttributeNameError:
                            errors.append(AttributeNameError({'type': 'Institution'}))

            # If the institution name matches the owner listed by the
            # script runner or this is the first institution,
            # add it as the owner of the topology
            if ((self.owner and str(institution_fields['name']).startswith(self.owner)) or not
                    primary_owner_imported):
                institution_fields['primary_owner'] = True
                primary_owner_imported = True

            # The add_institution method takes in multiple fields,
            # and returns the ID of the Institution object
            # For this script, we are not using the returned ID to
            # demonstrate using the get functions
            add_result = self.add_institution(**institution_fields)
            if not add_result['success']:
                errors.append(add_result['error'])

        # Collect the nodes from the Workbook
        nodes_sheet = workbook[WORKSHEET_LOCATIONS[template_version]['sheet_names']['nodes']]
        for row in nodes_sheet.iter_rows(
            min_row=WORKSHEET_LOCATIONS[template_version]['first_row']['nodes'],
            max_col=(max(WORKSHEET_LOCATIONS[template_version]['node_fields'].keys()) + 1),
            values_only=True,
        ):
            # Skipping over empty rows
            row = empty_row(row)
            if not row:
                continue

            # Fields in the database node
            node_fields = {}

            print('NODE: {}'.format(row[0]))  # RYAN

            # Collect the fields
            node_fields_map = WORKSHEET_LOCATIONS[template_version]['node_fields']
            for column, key in node_fields_map.items():
                if row[column]:
                    if ';' in str(row[column]):
                        node_fields[key] = [item.strip() for item in str(row[column]).split(';')]
                    else:
                        try:
                            if key == 'name' and row[column].strip() == '':
                                raise AttributeNameError
                            else:
                                node_fields[key] = row[column].strip()
                        except AttributeError:
                            # Likely not a string; just capture the
                            # value as-is
                            node_fields[key] = row[column]

            # Reset the owners field to collect the Institution IDs
            node_fields['owners'] = []
            owner_cells = [
                row[c] for c in node_fields_map.keys() if node_fields_map[c] == 'owners']
            for owner in concat_owners(*owner_cells):
                if not owner == 'None':
                    print('NODE OWNER: {}'.format(owner))  # RYAN
                    # The get function takes in key word arguments
                    # and returns a single object.
                    get_result = self.get_institution_by_name(
                        owner, 'node', node_fields.get('name', ''),
                    )
                    if not get_result['success']:
                        errors.append(get_result['error'])
                    node_fields['owners'].append(get_result['value'])

            add_result = self.add_node(**node_fields)
            if not add_result['success']:
                errors.append(add_result['error'])

        # Collect the links from the Workbook
        links_sheet = workbook[WORKSHEET_LOCATIONS[template_version]['sheet_names']['links']]
        for row in links_sheet.iter_rows(
            min_row=WORKSHEET_LOCATIONS[template_version]['first_row']['links'],
            max_col=(max(WORKSHEET_LOCATIONS[template_version]['link_fields'].keys()) + 1),
            values_only=True,
        ):
            # Skipping over empty rows
            row = empty_row(row)
            if not row:
                continue

            print('LINK: {}'.format(row[0]))

            # Fields in the database link
            link_fields = {}

            # Collect the fields
            link_fields_map = WORKSHEET_LOCATIONS[template_version]['link_fields']
            for column, key in link_fields_map.items():
                if row[column]:
                    if ';' in str(row[column]):
                        link_fields[key] = [item.strip() for item in str(row[column]).split(';')]
                    else:
                        try:
                            if key == 'name' and row[column].strip() == '':
                                raise AttributeNameError
                            link_fields[key] = row[column]
                        except AttributeError:
                            # Likely not a string; just capture the
                            # value as-is
                            link_fields[key] = row[column]

            endpoints = [row[c] for c in link_fields_map.keys() if link_fields_map[c] == 'nodes']
            # RYAN
            print('LINK ENDPOINT NODES: {} & {}'.format(*endpoints))

            node1_result = self.get_node_by_name(endpoints[0], link_fields.get('name', ''))
            if not node1_result['success']:
                errors.append(node1_result['error'])
            node2_result = self.get_node_by_name(endpoints[1], link_fields.get('name', ''))
            if not node2_result['success']:
                errors.append(node2_result['error'])
            link_fields['nodes'] = [
                node1_result['value'].id,
                node2_result['value'].id,
            ]

            # Reset the owners field to collect the Institution IDs
            link_fields['owners'] = []
            owner_cells = [
                row[c] for c in link_fields_map.keys() if link_fields_map[c] == 'owners']
            for owner in concat_owners(*owner_cells):
                if not owner == 'None':
                    print('LINK OWNER: {}'.format(owner))  # RYAN

                    get_result = self.get_institution_by_name(
                        owner, 'link', link_fields.get('name', ''),
                    )
                    if not get_result['success']:
                        errors.append(get_result['error'])
                    # The get function takes in key word arguments
                    # and returns a single object.
                    link_fields['owners'].append(get_result['value'])

            add_result = self.add_link(**link_fields)
            if not add_result['success']:
                errors.append(add_result['error'])

        return (self.manager, errors)
