import math
import torch
import torch.nn as nn
import openpyxl as xl
import openpyxl.utils as xlu


class ALGORITHMS:
    @staticmethod
    def linear_regression(data: list, EPOCH=500) -> (float, float):
        """
        Hybrid linear regression algorithm employing statistical analysis and machine learning.
        :param data: list, [x_data, y_data]
        :param EPOCH: int, Epochs for machine learning algorithm.
        :return: (float, float), slope / y-intercept

        Example:

        >>> lin_reg = algorithms_.linear_regression

        >>> x = [1, 2, 3]
        >>> y = [1, 2, 3]

        >>> data = [x, y]

        >>> m, b = lin_reg(data)
        >>> print(m, b)

        Returns: 1, 0
        """
        averages = []
        for dataset in data:
            temp = 0
            n = len(dataset)
            for datapoint in dataset:
                temp += datapoint / n

            averages.append(temp)

        variances = []

        if len(data) == len(averages):
            for i in range(len(averages)):
                n = len(data[i])
                temp = 0

                for j in range(n):
                    temp += pow(data[i][j] - averages[i], 2) / n

                variances.append(temp)

        covariance_xy = 0
        mu_x = averages[0]
        mu_y = averages[1]
        x = data[0]
        y = data[1]
        n = len(data[0])

        for i in range(n):
            covariance_xy += (x[i] - mu_x) * (y[i] - mu_y) / n

        correlation = covariance_xy / math.sqrt(variances[0] * variances[1])

        m = correlation * math.sqrt(variances[1] / variances[0])

        x = []
        y = []
        for i in range(len(data[0])):
            x.append([data[0][1]])
            y.append([data[1][1]])

        x = torch.Tensor(x)
        y = torch.Tensor(y)

        b = nn.Parameter(torch.randn(1, 1))
        M = torch.Tensor([m])
        loss_f = torch.nn.MSELoss(size_average=False)
        optimizer = torch.optim.SGD([b], lr=0.001)

        epoch = 0
        while epoch < EPOCH:
            pred_y = M * x + b
            optimizer.zero_grad()
            loss = loss_f(pred_y, y)
            loss.backward()
            optimizer.step()

            epoch += 1

        return m, b.item()

    @staticmethod
    def correlation_coefficient_matrix(data: list) -> list:
        """
        Correlation coefficient matrix
        :param data: list
        :return: list

        Example:

        >>> f_x = algorithms_.correlation_coefficient_matrix
        >>> V_1 = [1, 1, 1]
        >>> V_2 = [1, 1, 1]

        >>> V_3 = f_x(V_1, V_2)
        >>> print(V_3)

        Returns: [[1, 1]
                  [1, 1]]
        """
        n_cols = len(data)
        n = len(data[0])
        row = [0 for i in range(n_cols)]
        matrix = [row for i in range(n_cols)]

        for i in range(n_cols):
            for j in range(n_cols):
                if i != j:
                    cov = 0
                    # means start
                    means = []
                    for dataset in data:
                        temp = 0
                        n = len(dataset)
                        for datapoint in dataset:
                            temp += datapoint / n

                        means.append(temp)
                    # means end
                    # variance start
                    variances = []

                    if len(data) == len(means):
                        for l in range(len(means)):
                            n = len(data[l])
                            temp = 0

                            for m in range(n):
                                temp += pow(data[l][m] - means[l], 2) / n

                            variances.append(temp)
                    # variance end

                    for k in range(n):
                        cov += (data[i][k] - means[0])*(data[j][k] - means[1])/n

                    matrix[i][j] += cov/pow(variances[0]*variances[1], 0.5)

                if i == j:
                    matrix[i][j] = 0

        return matrix

    def multiple_linear_regression(self, data: list, depth: int) -> list:
        pred = []
        for d in range(depth):
            # averages for data
            averages = []
            for dataset in data:
                temp = 0
                n = len(dataset)
                for datapoint in dataset:
                    temp += datapoint / n

                averages.append(temp)
            cols = len(data)

            # variances for data
            variances = []

            if len(data) == len(averages):
                for i in range(len(averages)):
                    n = len(data[i])
                    temp = 0

                    for j in range(n):
                        temp += pow(data[i][j] - averages[i], 2) / n

                    variances.append(temp)

            corr_coeff_matrix = self.correlation_coefficient_matrix(data)
            A = [0 for i in range(cols)]
            p_matrix = [A for i in range(cols)]

            # p_matrix -> slope matrix
            S = [0 for c in range(cols)]
            for i in range(cols):
                for j in range(cols):
                    if corr_coeff_matrix[i][j] > 0.7 and i != j:
                        p_matrix[i][j] += corr_coeff_matrix[i][j]*math.sqrt(variances[i]/variances[j])

                    if i != j:
                        S[i] += p_matrix[i][j]


        return pred

    @staticmethod
    def integral(func, func_args: list, interval: list) -> float:
        """
        Returns the area under a defined continuous function on an interval.
        :param func: callable function with x as first param.
        :param func_args: list, arguments for callable function that returns float.
        :param interval: interval on which the geometric area is to be calculated.
        :return: float

        Example:

        >>> def f_x(X):
        >>>     return 2

        >>> I = algorithms_.integral

        >>> Area = I(f_x, [2, 6])
        >>> print(Area)

        Returns: 8
        """
        dx = 0.0001
        area = 0
        x = interval[0]
        while x < interval[1]:
            area += func(x, func_args) * dx
            x += dx

        return area


class VECTOR:
    @staticmethod
    def vector_add(vector_1: list, vector_2: list, add: bool = True) -> list:
        """
        Returns the sum or difference between two vectors.
        :param vector_1: list
        :param vector_2: list
        :param add: bool, Default: True
        :return: list

        Example:

        >>> f_x = vector_.vector_cross
        >>> V_1 = [1, 2, 3]
        >>> V_2 = [1, 1, 1]

        >>> V_3 = f_x(V_1, V_2)
        >>> print(V_3)

        Returns: [2, 3, 4]
        """
        new_vector = []
        for element_x, element_y in vector_1, vector_2:
            if add:
                new_vector.append(element_x + element_y)
            if not add:
                new_vector.append(element_x - element_y)

        return new_vector

    @staticmethod
    def vector_dot(vector_1: list, vector_2: list) -> float:
        """
        Returns the dot product between two vectors.
        :param vector_1: list
        :param vector_2: list
        :return: float

        Example:

        >>> f_x = vector_.vector_cross
        >>> V_1 = [1, 2, 3]
        >>> V_2 = [3, 2, 1]

        >>> scalar = f_x(V_1, V_2)
        >>> print(scalar)

        Returns: 10
        """
        dot_sum = 0
        for element_x, element_y in vector_1, vector_2:
            dot_sum += element_x * element_y

        return dot_sum

    @staticmethod
    def vector_cross(vector_1: list, vector_2: list) -> list:
        """
        Returns the cross product between two 3-dimensional vectors.
        :param vector_1: list
        :param vector_2: list
        :return: list

        Example:

        >>> cross_prod = vector_.vector_cross
        >>> V_1 = [1, 0, 0]
        >>> V_2 = [0, 1, 0]

        >>> V_3 = cross_prod(V_1, V_2)
        >>> print(V_3)

        Returns: [0, 0, 1]
        """
        a = vector_1[1] * vector_2[2] - vector_1[2] * vector_2[1]
        b = vector_1[0] * vector_2[2] - vector_1[2] * vector_2[0]
        c = vector_1[0] * vector_2[1] - vector_1[1] * vector_2[0]

        return [a, b, c]


class SPREADSHEET:
    @staticmethod
    def accepted_dtypes() -> list:
        """
        Returns a list of accepted dtypes for scripting table import.
        :return: list, [float, int, str]
        """
        return [float, int, str]

    def import_excel_table(self, cols: int, rows: int, dtype: list, *, filename: str):
        """
        Imports a table from Excel. Table must start at cell 'A2'.
        :param cols: int, Number of columns in the Excel table.
        :param rows: int, Number of rows in the Excel table.
        :param dtype: list, list containing dtypes describing each column.
        :param filename: str, Name of the Excel file.
        :return: list, data is formatted by columns.
        """
        ws = xl.load_workbook(filename).active
        _columns = [xlu.get_column_letter(i + 1) for i in range(cols)]
        _rows = [i + 2 for i in range(rows)]

        data = []
        for row in _rows:
            temp = []
            skip_row = False
            column_num = 0

            for col in _columns:
                cell_value = ws[f'{col}{row}'].value
                temp.append(cell_value)

                if type(cell_value) is not dtype[column_num]:
                    skip_row = True

                column_num += 1

            if not skip_row:
                data.append(temp)

            formatted_data = self.format_to_cols(data)
            return formatted_data

    @staticmethod
    def format_to_cols(data: list) -> list:
        """
        Formats data list from row-based to column-based.
        :param data: list, data.
        :return: list

        Example:

        >>> data_ = [[1, 2, 3], [2, 4, 6], [3, 6, 9], [4, 8, 12]]
        >>> data_ = spreadsheet_.format_to_cols(data_)

        returns: [[1, 2, 3, 4], [2, 4, 6, 8], [3, 6, 9, 12]]

        """
        cols = len(data[0])
        formatted_data = [[] for i in range(cols)]

        for row in data:
            for i in range(cols):
                formatted_data[i].append(row[i])

        return formatted_data

    @staticmethod
    def format_to_rows(data: list) -> list:
        """
        Formats data from column-based to row-based.
        :param data: list, data.
        :return: list

        Example:

        >>> data_ = [[1, 2, 3, 4], [2, 4, 6, 8], [3, 6, 9, 12]]
        >>> data_ = spreadsheet_.format_to_cols(data_)

        returns: [[1, 2, 3], [2, 4, 6], [3, 6, 9], [4, 8, 12]]

        """
        rows = len(data[0])
        formatted_data = [[] for i in range(rows)]
        for col in data:
            for r in range(rows):
                formatted_data[r].append(col[r])

        return formatted_data


algorithms_, vector_, spreadsheet_ = ALGORITHMS(), VECTOR(), SPREADSHEET()
