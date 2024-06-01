import openpyxl as xlx
import xlrd
import odsparsator.odsparsator as ods

class xlFileErrorSheetnameError(Exception):
    pass


class xlFile:
    '''
    Convenience class for supporting .xlsx and .xls files
    without caring about details.
    '''

    def __init__(self, filename):
        self.wb = None

    @classmethod
    def load_spreadsheet(cls, filename):
        try:                    # Most of the time we will get .xlsx files, so let's try that
            return xlsxFile(filename)
        except xlx.utils.exceptions.InvalidFileException:
            try:            # Try open an .xls file
                return xlsFile(filename) # Need exception handling here too?
            except xlrd.biffh.XLRDError:
                return odsFile(filename)


class xlsxFile(xlFile):
    def __init__(self, filename):
        self.wb = xlx.load_workbook(filename, data_only=True)

    def sheet_names(self):
        return self.wb.sheetnames


    def sheet(self, sheetname):
        try:
            return xlsxSheet(self.wb[sheetname])
        except KeyError:
            raise xlFileErrorSheetnameError()


class xlsxSheet:
    def __init__(self, sheet):
        self.sh = sheet

    def n_columns(self):
        return self.sh.max_column

    def rows(self):
        return self.sh.values



class xlsFile(xlFile):
    def __init__(self, filename):
        self.wb = xlrd.open_workbook(filename)

    def sheet_names(self):
        return self.wb.sheet_names()

    def sheet(self, sheetname):
        return xlsSheet(self.wb.sheet_by_name(sheetname))


class xlsSheet:
    def __init__(self, sheet):
        self.sh = sheet

    def n_columns(self):
        return self.sh.ncols

    def rows(self):
        for r in self.sh.get_rows():
            yield map(lambda c: c.value, r)

    
class odsFile(xlFile):
    def __init__(self, filename):
        wb = ods.ods_to_python(filename, export_minimal=True)
        if type(wb) == dict:
            self.wb = wb['body']
        else:
            self.wb = wb        # "A list of tabs", according to the documentation.
            
    def sheet_names(self):
        return list(map(lambda tab: tab['name'], self.wb))

    def sheet(self, sheetname):
        sheet = None
        for sheet_data in self.wb:
            if sheet_data['name'] == sheetname:
                return odsSheet(sheet_data['table'])
        raise xlFileErrorSheetnameError() # If we use the wrong name
            

class odsSheet:
    def __init__(self, sheet):
        self.sh = sheet

    def n_columns(self):
        return max(map(len, self.sh)) # Find the row with the most elements

    def rows(self):
        return self.sh
