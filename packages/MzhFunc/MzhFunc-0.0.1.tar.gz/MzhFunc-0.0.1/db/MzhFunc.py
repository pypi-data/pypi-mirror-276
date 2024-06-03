from subprocess import PIPE, Popen
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.header import Header
import requests
import pandas as pd
import xlrd
import openpyxl
import os
"""
2099cd35d78f6431
5f9600081ec1b0ea
6dfe56f9a32582af
e0dd19d1803a51d7
da4710f18a9a19ed
e51e17c80da377b4
732a3033a742f248
77fb64e1c22d6ceb

cd pip && python setup.py sdist
twine upload dist/*
pypi-AgEIcHlwaS5vcmcCJDRmMGRhNGJiLTM4ZDktNDlkNi1iMjc1LWExNmU3YmIwYmFkMwACKlszLCJjMjQ1NzNhNy01ZjQzLTQ2MDktODFmZi04MjQ5NjY4MmY0MWUiXQAABiDkEBiGZim6aIww1x2eC03oY0MKJdNumwZ8d5Gpd9F1pg
"""


def cmd_utf(order):
    proc = Popen(order, stdin=None, stdout=PIPE, stderr=PIPE, shell=True)
    outinfo, errinfo = proc.communicate()
    outinfo = outinfo.decode('utf-8')
    errinfo = errinfo.decode('utf-8')
    return outinfo, errinfo


def cmd_gbk(order):
    proc = Popen(order, stdin=None, stdout=PIPE, stderr=PIPE, shell=True)
    outinfo, errinfo = proc.communicate()
    outinfo = outinfo.decode('gbk')
    errinfo = errinfo.decode('gbk')
    return outinfo, errinfo


def py_run(path):
    try:
        print(cmd_gbk(rf"cd /d {os.path.dirname(path)} && python {os.path.basename(path)}")[0])
        return
    except:
        pass
    print(cmd_utf(rf"cd /d {os.path.dirname(path)} && python {os.path.basename(path)}")[0])


def pushplus(title, content, token="cf8736250267472e954737c221b33d23"):
    url = fr"https://www.pushplus.plus/send?token={token}&title={title}&content={content}&template=html"
    payload = {}
    files = {}
    headers = {
        'User-Agent': 'Apifox/1.0.0 (https://apifox.com)'
    }
    requests.request("GET", url, headers=headers, data=payload, files=files)


def send_email(name_list, fpath, text):
    # 生成连结对象,参数分别是邮件服务器和端口号
    con = smtplib.SMTP_SSL('smtp.qq.com', 465)
    # 使用用户名和密码登录,这里密码以星号隐藏了
    con.login('191891173@qq.com', 'tvthacuvztgicaib')
    # 生成一个邮件对象，由于邮件包含文本、图片、HTML、附件等内容，
    # 所以这里用MIMEMultipart()生成邮件对象，以支持多种数据格式
    mail_obj = MIMEMultipart()
    # 生成邮件表头的内容
    mail_header = Header(text, 'utf-8').encode()
    # 主题
    mail_obj['Subject'] = mail_header
    # 发送者邮箱
    mail_obj['From'] = '191891173@qq.com <191891173@qq.com>'
    # 接收者邮箱
    mail_obj['To'] = '我'
    # 添加邮件正文
    mail_text = MIMEText('888', 'plain', 'utf-8')
    mail_obj.attach(mail_text)
    for path in fpath:
        if path.endswith('.txt'):
            # 添加txt附件
            with open(path, 'rb') as f:
                txt = f.read()
                txt = MIMEText(txt, 'base64', 'utf-8')
                txt["Content-Type"] = 'application/octet-stream'
                txt["Content-Disposition"] = 'attachment; filename="I.txt"'
                mail_obj.attach(txt)
        if path.endswith('.xlsx'):
            # 添加Excel附件
            with open(path, 'rb') as f:
                Excel = f.read()
                Excel = MIMEText(Excel, 'base64', 'utf-8')
                Excel["Content-Type"] = 'application/octet-stream'
                Excel["Content-Disposition"] = 'attachment; filename="ove.xlsx"'
                mail_obj.attach(Excel)
        if path.endswith('.zip'):
            # 添加Zip附件
            with open(path, 'rb') as f:
                Zip = f.read()
                Zip = MIMEText(Zip, 'base64', 'utf-8')
                Zip["Content-Type"] = 'application/octet-stream'
                Zip["Content-Disposition"] = 'attachment; filename="class.rar"'
                mail_obj.attach(Zip)
        if path.endswith('.png'):
            # 添加图片附件
            with open(path, 'rb') as f:
                img2 = f.read()
                img_2 = MIMEImage(img2)
                # 指定图片类型与文件名，以下语句设置图片文件以附件形式加到邮件中
                img_2['Content-Disposition'] = 'attachment;filename="flower.png"'
                # 加入到邮件中
                mail_obj.attach(img_2)
        if path.endswith('.docx'):
            # 添加word附件
            with open(path, 'rb') as f:
                doc = f.read()
                # 以数据流的形式读入文件
                doc = MIMEText(doc, 'base64', 'utf-8')
                # 以下语句设置文件以附件形式加到邮件中
                doc['Content-Disposition'] = 'attachment;filename="test.docx"'
                # 加入到邮件中
                mail_obj.attach(doc)

    # 发送邮件
    con.sendmail('191891173@qq.com', name_list, mail_obj.as_string())
    # 断开连结
    con.quit()
    print('发送邮件成功...')


def read_excel(path, sheet_name):
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_name(sheet_name)
    data_list = [sheet.row_values(rowx=i) for i in range(sheet.nrows)]
    return data_list


def writelist_toExcel(ls, savename, loadname=None, sheet_name=None):
    if loadname == None:
        book = openpyxl.Workbook()
        sh = book.active
        for row in ls:
            sh.append(row)
        book.save(savename)
    else:
        wb = openpyxl.load_workbook(loadname)
        sheet = wb[sheet_name]
        for row in ls:
            sheet.append(row)
        wb.save(savename)


def pd_read_excel(path, sheet_name):
    datalist = pd.read_excel(path, sheet_name, header=None, index_col=None)
    LS = []
    for j in range(len(datalist)):
        ls = []
        for i in datalist.columns:
            ls.append(datalist.iloc[j][i])
        LS.append(ls)
    return LS


def pd_writelist_toExcel(ls, savename, sheet_name):
    if not os.path.exists(savename):
        book = openpyxl.Workbook()
        sh = book.active
        sh.title = sheet_name
        book.save(savename)
    df = pd.DataFrame(ls)
    with pd.ExcelWriter(savename, datetime_format="YYYY-MM-DD") as writer:
        df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)


def pd_writelist_appendExcel(ls, savename, sheet_name):
    data = pd_read_excel(savename, sheet_name)
    ls = data + ls
    df = pd.DataFrame(ls)
    with pd.ExcelWriter(savename, datetime_format="YYYY-MM-DD") as writer:
        df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)