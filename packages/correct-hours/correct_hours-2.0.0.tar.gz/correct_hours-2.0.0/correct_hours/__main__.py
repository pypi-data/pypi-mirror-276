import logging
from openpyxl import load_workbook
from pathlib import Path

from correct_hours.report_processors.myob import MyobReportProcessor
from correct_hours.report_processors.xero import XeroReportProcessor
from correct_hours.types import RateFileNotFound, InvalidReportType, RATES_FILENAME, OUTPUT_FOLDER, \
    HOURS_NEW_FILE_PREFIX

import tkinter as tk
from tkinter import filedialog, ttk, font


def get_new_file_name(filepath: Path) -> str:
    path = Path(filepath)
    # Example: ~/correct-hours/examples/xero/output/copy_barney-stinson.xlsx
    return f"{path.parent.absolute()}/{OUTPUT_FOLDER}/{HOURS_NEW_FILE_PREFIX}{path.name}"


def should_ignore_file(file: Path) -> bool:
    filename = file.name
    return (
            str.startswith(filename, "~") or
            filename == RATES_FILENAME
    )


def process_directory(directory: str, timesheet_type: str) -> None:
    # create output folder
    Path(f"{directory}/{OUTPUT_FOLDER}").mkdir(parents=True, exist_ok=True)
    # iterate through files in input directory
    files = Path(directory).glob('*')
    for f in files:
        if f.is_file():
            if not should_ignore_file(f):
                hours_filepath = f.absolute()
                print(f"Processing file {hours_filepath}...")
                workbook = load_workbook(filename=hours_filepath)
                # load rates workbook
                rates_filepath = f"{directory}/{RATES_FILENAME}"
                if timesheet_type == 'xero':
                    try:
                        rates_workbook = load_workbook(filename=rates_filepath)
                    except FileNotFoundError:
                        raise RateFileNotFound(rates_filepath)
                    processor = XeroReportProcessor(workbook, rates_workbook)
                elif timesheet_type == 'myob':
                    processor = MyobReportProcessor(workbook)
                else:
                    raise InvalidReportType(timesheet_type)
                processor.process()
                hours_new_filename = get_new_file_name(hours_filepath)
                workbook.save(filename=hours_new_filename)
                print(f"Finished processing file. Created file {hours_new_filename}.")


def start():
    # Disable the submit button to prevent multiple submissions
    start_button.config(state="disabled")

    # Display loading animation
    result_label.config(text="Loading...")
    result_label.update()

    try:
        directory = directory_input.get()
        process_directory(directory=directory, timesheet_type=timesheet_type.get())
        result_label.config(
            text=f'Timesheets have been corrected. They have been saved in "{directory}/{OUTPUT_FOLDER}".')
    except Exception as e:
        error_message = "Failed to process directory"
        logging.error(error_message, exc_info=e)
        result_label.config(text=error_message)

    result_label.update()

    # Enable the submit button after loading
    start_button.config(state="normal")


def browse_directory():
    directory_path = filedialog.askdirectory()
    if directory_path:
        directory_input.delete(0, tk.END)
        directory_input.insert(tk.END, directory_path)


# Create the main application window
app = tk.Tk()
app.title("Overtime in timesheets")

# italic font
default_font = font.nametofont('TkDefaultFont')
italic_font = default_font.copy()
italic_font.config(slant='italic')

# Create and place widgets

# report type field
timesheet_type_label = tk.Label(app, text="Select timesheet type:")
timesheet_type_label.grid(row=0, column=0, padx=10, pady=0, sticky=tk.W)

timesheet_type = tk.StringVar()
timesheet_type_combo = ttk.Combobox(app, textvariable=timesheet_type, state="readonly")
timesheet_type_combo['values'] = ('xero',)
timesheet_type_combo.current(0)
timesheet_type_combo.grid(row=0, column=1, padx=10, pady=0, sticky=tk.W)

# directory field
directory_label = tk.Label(app, text="Select a directory:")
directory_label.grid(row=1, column=0, padx=10, pady=0, sticky=tk.W)

directory_input = tk.Entry(app, width=100)
directory_input.grid(row=1, column=1, padx=10, pady=0, sticky=tk.W)

directory_input_description = tk.Label(app, text="This is where the timesheets to be processed are located.", font=italic_font)
directory_input_description.grid(row=2, column=1, padx=10, pady=0, sticky=tk.W)

# browse button
browse_button = tk.Button(app, text="Browse", command=browse_directory)
browse_button.grid(row=1, column=2, padx=10, pady=0)

# submit button
start_button = tk.Button(app, text="Start", command=start)
start_button.grid(row=3, column=0, columnspan=3, padx=10, pady=10)

# result label
result_label = tk.Label(app, text="")
result_label.grid(row=4, column=0, columnspan=3, padx=10, pady=10)

# Run the application
app.mainloop()
