from datetime import datetime
from typing import Tuple

from openpyxl import Workbook
from openpyxl.styles import Font

from correct_hours.report_processors.xero.rate_processor import XeroRateProcessor
from correct_hours.types import InvalidRateLabel
from correct_hours.utils import get_col_number, trim_rate_label

ROW_OFFSET = 6


class XeroReportProcessor:

    def __init__(self, hours_workbook: Workbook, rates_workbook: Workbook) -> None:
        self.hours_workbook = hours_workbook
        self.rates_workbook = rates_workbook
        self.hours_original_sheet = self.hours_workbook.active
        self.hours_new_sheet = self.hours_workbook.copy_worksheet(from_worksheet=self.hours_original_sheet)

    def add_new_column_headings(self) -> None:
        # new column headings
        self.hours_new_sheet.cell(5, 15, "New total").font = Font(bold=True)
        self.hours_new_sheet.cell(5, 16, "New total per day").font = Font(bold=True)
        self.hours_new_sheet.cell(5, 17, "Days worked").font = Font(bold=True)
        self.hours_new_sheet.cell(5, 18, "Rates").font = Font(bold=True)

    def add_up_hours_of_the_same_day(self, current_date: datetime, current_employee: str, row_start: int) -> Tuple[int, int]:
        rows_added_count = 0
        total_hours = 0
        for row in self.hours_new_sheet.iter_rows(min_row=row_start, min_col=1, max_col=14, values_only=True):
            date = row[0]
            employee = row[1]
            if date != current_date or employee != current_employee:
                break
            total = row[13]
            total_hours += total
            rows_added_count += 1
        return rows_added_count, total_hours

    def correct_hours(self, row_start: int, row_end: int, overtime: int) -> None:
        time_left = overtime
        for col_idx in range(13, 7, -1):
            for row_idx in range(row_end, row_start - 1, -1):
                value = self.hours_new_sheet.cell(row_idx, col_idx).value
                if value <= 0:
                    continue
                corrected_value = value - time_left
                if corrected_value < 0:
                    time_left = corrected_value * -1
                    corrected_value = 0
                    self.hours_new_sheet.cell(row_idx, col_idx, corrected_value).font = Font(bold=True, color="FF0000")
                else:
                    self.hours_new_sheet.cell(row_idx, col_idx, corrected_value).font = Font(bold=True, color="FF0000")
                    return

    def process_rows(self) -> None:
        # load hour rates to keep them handy
        rate_processor = XeroRateProcessor(self.rates_workbook)
        rate_processor.process()
        # start processing hour rows
        rows_processed_count = 0
        for row_idx, row in enumerate(
                self.hours_new_sheet.iter_rows(min_row=1, min_col=1, max_col=14, values_only=True)
        ):
            row_number = row_idx + 1
            # validating row
            if not row[0]:
                # skip empty row
                rows_processed_count += 1
                continue
            if row_number < ROW_OFFSET:
                # skip the header rows
                rows_processed_count += 1
                continue
            rate_label = row[3]
            rate_label_trimmed = trim_rate_label(rate_label)
            if not rate_label_trimmed:
                raise InvalidRateLabel(rate_label, row_number)

            # add new total
            self.hours_new_sheet.cell(row_number, 15, f"=SUM(G{row_number}:M{row_number})")
            # apply rate
            date = row[0]
            rate = rate_processor.get_rate(rate_label_trimmed, date)
            rate_cell = self.hours_new_sheet.cell(row_number, 18, f"=O{row_number}*{rate}")
            rate_cell.style = 'Currency'
            if row_number <= rows_processed_count:
                # skip the rows of the same day
                continue
            employee = row[1]
            rows_added_count, total_hours = self.add_up_hours_of_the_same_day(date, employee, row_number)
            # last row of the same day
            last_row_number = (row_number + rows_added_count) - 1
            did_overtime = False
            if total_hours > 38:
                overtime = total_hours - 38
                self.correct_hours(row_number, last_row_number, overtime)
                did_overtime = True
            # new total per day
            new_total_of_hours = self.hours_new_sheet.cell(
                last_row_number, 16, f"=SUM(O{row_number}:O{last_row_number})"
            )
            if did_overtime:
                new_total_of_hours.font = Font(bold=True, color="FF0000")
            # number of days worked
            days_worked_cell = self.hours_new_sheet.cell(last_row_number, 17, f"=P{last_row_number}/7.6")
            days_worked_cell.style = 'Comma [0]'

            rows_processed_count += rows_added_count
        # add total cells at the bottom
        bottom_row = rows_processed_count + 1
        self.hours_new_sheet.cell(
            bottom_row,
            get_col_number("Q"),
            f"=SUM(Q{ROW_OFFSET}:Q{rows_processed_count})"
        ).style = "Comma [0]"
        self.hours_new_sheet.cell(
            bottom_row,
            get_col_number("R"),
            f"=SUM(R{ROW_OFFSET}:R{rows_processed_count})"
        ).style = "Currency"

    def process(self) -> None:
        self.add_new_column_headings()
        self.process_rows()
